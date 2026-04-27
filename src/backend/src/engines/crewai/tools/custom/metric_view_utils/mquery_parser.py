"""MQuery Parser — extract table structure from transpiled SQL."""
from __future__ import annotations

import json
import re

from .constants import (
    RE_AGG_COL,
    RE_CALC_COL,
    RE_CASE_AGG,
    RE_COALESCE_AGG,
    RE_FROM_CLAUSE,
    RE_GROUP_BY,
    RE_LEFT_JOIN,
)
from .data_classes import TableInfo


class MQueryParser:
    """Parse MQuery conversion report (JSON or Excel) and extract table structure per table."""

    def parse_json(self, json_path: str | list[dict]) -> dict[str, TableInfo]:
        """Parse mquery_transpilation JSON — accepts file path or raw list."""
        if isinstance(json_path, list):
            entries = json_path
        else:
            with open(json_path) as f:
                entries = json.load(f)
        tables: dict[str, TableInfo] = {}
        for entry in entries:
            table_name = entry.get('table_name', '')
            sql = entry.get('transpiled_sql', '')
            status = entry.get('validation_passed', '')
            if not table_name or not sql:
                continue
            if not isinstance(status, str) or not status.startswith('Yes'):
                if not ('SUM(' in sql.upper() and 'GROUP BY' in sql.upper()):
                    continue
            info = self._parse_sql(table_name, sql)
            info.raw_transpiled_sql = sql
            tables[table_name] = info
        return tables

    def parse(self, xlsx_path: str) -> dict[str, TableInfo]:
        """Parse MQuery_Conversion_Report.xlsx (requires openpyxl)."""
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb['MQuery Conversion Report']
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_idx = {h: i + 1 for i, h in enumerate(headers)}
        tables: dict[str, TableInfo] = {}
        for r in range(2, ws.max_row + 1):
            table_name = ws.cell(row=r, column=col_idx['Table Name']).value
            status = str(ws.cell(row=r, column=col_idx['Validation Passed']).value or '')
            sql = str(ws.cell(row=r, column=col_idx['Transpiled SQL']).value or '')
            if not table_name or not sql:
                continue
            if not isinstance(status, str) or not status.startswith('Yes'):
                if not ('SUM(' in sql.upper() and 'GROUP BY' in sql.upper()):
                    continue
            info = self._parse_sql(table_name, sql)
            tables[table_name] = info
        return tables

    def _parse_sql(self, table_name: str, sql: str) -> TableInfo:
        """Parse transpiled SQL to extract structure."""
        is_cte = sql.strip().upper().startswith('CREATE') and 'WITH ' in sql.upper()
        main_sql = self._extract_final_select(sql) if is_cte else sql

        # Extract source table from FROM clause
        source_table = ''
        all_from = list(RE_FROM_CLAUSE.finditer(sql))
        main_from = list(RE_FROM_CLAUSE.finditer(main_sql))
        if main_from:
            for m in main_from:
                if '.' in m.group(1):
                    source_table = m.group(1)
                    break
        if not source_table and all_from:
            for m in all_from:
                if '.' in m.group(1):
                    source_table = m.group(1)
                    break

        # Extract JOIN tables
        dim_source_tables: dict[str, str] = {}
        for m in RE_LEFT_JOIN.finditer(sql):
            join_table = m.group(1)
            join_alias = m.group(2)
            if '.' in join_table:
                dim_source_tables[join_alias] = join_table

        # Collect ALL SQL aliases for filter rewriting
        all_sql_aliases: set[str] = set(dim_source_tables.keys())
        for m_from in re.finditer(r'\bFROM\s+[\w.]+\s+(?:AS\s+)?(\w+)', sql, re.IGNORECASE):
            all_sql_aliases.add(m_from.group(1))
        for m_join in re.finditer(r'\bJOIN\s+[\w.]+\s+(?:AS\s+)?(\w+)\s+ON\b', sql, re.IGNORECASE):
            all_sql_aliases.add(m_join.group(1))

        # Extract aggregate columns
        agg_scan_sql = main_sql
        if is_cte and not RE_AGG_COL.search(main_sql):
            agg_scan_sql = sql

        agg_cols = []
        seen_agg_names: set[str] = set()
        for m in RE_AGG_COL.finditer(agg_scan_sql):
            source_col = m.group(1)
            alias = m.group(2)
            if alias not in seen_agg_names:
                agg_cols.append({'name': alias, 'source_col': source_col})
                seen_agg_names.add(alias)

        for m in RE_COALESCE_AGG.finditer(agg_scan_sql):
            alias = m.group(3)
            if alias not in seen_agg_names:
                expr = f'COALESCE(SUM(source.{m.group(1)}), 0) - COALESCE(SUM(source.{m.group(2)}), 0)'
                agg_cols.append({'name': alias, 'source_col': alias, 'expr': expr})
                seen_agg_names.add(alias)

        for m in RE_CASE_AGG.finditer(agg_scan_sql):
            alias = m.group(1)
            if alias not in seen_agg_names:
                full_match = m.group(0)
                agg_cols.append({'name': alias, 'source_col': alias, 'expr': full_match.rsplit(' AS ', 1)[0].strip()})
                seen_agg_names.add(alias)

        # Extract GROUP BY
        group_cols = []
        gb_match = RE_GROUP_BY.search(main_sql)
        if gb_match:
            gb_text = gb_match.group(1).strip()
            if gb_text.upper() == 'ALL':
                group_cols = self._infer_group_by_all(main_sql, seen_agg_names)
            else:
                for col in gb_text.split(','):
                    col = col.strip().rstrip(',')
                    if col:
                        if '.' in col:
                            col = col.split('.')[-1]
                        group_cols.append(col)

        # Clean group cols
        _valid_ident = re.compile(r'^[a-zA-Z_]\w*$')
        seen_dims: set[str] = set()
        clean_cols: list[str] = []
        for c in group_cols:
            if _valid_ident.match(c) and c not in seen_dims:
                seen_dims.add(c)
                clean_cols.append(c)
        _base_col_names = set(clean_cols)
        filtered_cols = []
        for c in clean_cols:
            m_suffix = re.match(r'^(.+?)(\d)$', c)
            if m_suffix and m_suffix.group(1) in _base_col_names:
                continue
            filtered_cols.append(c)
        group_cols = filtered_cols

        # Extract WHERE conditions
        static_filters: list[str] = []
        where_match = re.search(
            r'\bWHERE\b\s+(.*?)(?:\bGROUP\s+BY\b|\bORDER\s+BY\b|\bHAVING\b|$)',
            main_sql, re.IGNORECASE | re.DOTALL,
        )
        if where_match:
            where_text = where_match.group(1).strip()
            conditions = self._split_where_conditions(where_text)
            for cond in conditions:
                cond = cond.strip()
                if not cond:
                    continue
                if re.search(r'[:$]\{?\w+\}?', cond):
                    continue
                if re.search(r'\bCASE\b', cond, re.IGNORECASE):
                    case_count = len(re.findall(r'\bCASE\b', cond, re.IGNORECASE))
                    end_count = len(re.findall(r'\bEND\b', cond, re.IGNORECASE))
                    if case_count > end_count:
                        continue
                single_quotes = cond.count("'")
                if single_quotes % 2 != 0:
                    continue
                if not re.search(r'[a-zA-Z_]\w*\s*[=<>!]', cond) and 'IN' not in cond.upper():
                    continue
                for alias in all_sql_aliases:
                    cond = re.sub(rf'\b{re.escape(alias)}\.(\w+)', r'source.\1', cond)
                cond = re.sub(r'\b[a-z]\d*\.(\w+)', r'source.\1', cond)
                static_filters.append(cond)

        # Extract calculated columns
        calc_cols = []
        in_calc_section = False
        for line in sql.split('\n'):
            if '-- Calculated Columns' in line:
                in_calc_section = True
                continue
            if in_calc_section:
                m = RE_CALC_COL.match(line)
                if m:
                    expr = m.group(1).strip().rstrip(',')
                    name = m.group(2)
                    calc_cols.append({'name': name, 'expr': expr})
                elif line.strip().startswith('FROM') or line.strip().startswith('WHERE'):
                    in_calc_section = False

        is_fact = len(agg_cols) > 0
        return TableInfo(
            table_name=table_name,
            source_table=source_table,
            aggregate_columns=agg_cols,
            group_by_columns=group_cols,
            calculated_columns=calc_cols,
            is_fact=is_fact,
            full_sql=sql,
            dim_source_tables=dim_source_tables,
            static_filters=static_filters,
        )

    def _extract_final_select(self, sql: str) -> str:
        """For CTE queries, extract the first top-level SELECT after CTEs."""
        lines = sql.split('\n')
        first_select_idx = -1
        in_paren_depth = 0
        for i, line in enumerate(lines):
            stripped = line.strip()
            in_paren_depth += stripped.count('(') - stripped.count(')')
            if stripped.upper().startswith('SELECT') and in_paren_depth <= 0:
                first_select_idx = i
                break
        if first_select_idx > 0:
            return '\n'.join(lines[first_select_idx:])
        return sql

    @staticmethod
    def _split_where_conditions(where_text: str) -> list[str]:
        """Split WHERE clause on top-level AND, respecting parentheses."""
        conditions: list[str] = []
        depth = 0
        current: list[str] = []
        for token in re.split(r'\b(AND)\b', where_text, flags=re.IGNORECASE):
            if token.upper() == 'AND' and depth == 0:
                conditions.append(' '.join(current).strip())
                current = []
            else:
                depth += token.count('(') - token.count(')')
                current.append(token)
        if current:
            conditions.append(' '.join(current).strip())
        return conditions

    def _infer_group_by_all(self, sql: str, agg_names: set[str]) -> list[str]:
        """For GROUP BY ALL, infer dimensions from SELECT columns that aren't aggregates."""
        union_match = re.search(r'\bUNION\s+(?:ALL\s+)?SELECT\b', sql, re.IGNORECASE)
        if union_match:
            sql = sql[:union_match.start()]
        cols = []
        select_match = re.search(r'SELECT\s+([\s\S]+?)\s+FROM\s+', sql, re.IGNORECASE)
        if not select_match:
            return cols
        select_block = select_match.group(1)
        column_exprs = self._split_select_columns(select_block)
        for expr in column_exprs:
            expr = expr.strip()
            if not expr:
                continue
            if re.match(r'(?:sum|count|avg|min|max)\s*\(', expr, re.IGNORECASE):
                continue
            if 'COALESCE(SUM' in expr.upper():
                continue
            if '-- Calculated' in expr:
                continue
            as_match = re.search(r'\bAS\s+`?(\w+)`?\s*$', expr, re.IGNORECASE)
            if as_match:
                name = as_match.group(1)
            else:
                name = expr.split('.')[-1].strip().rstrip(',')
            if name and name not in agg_names and re.match(r'^[a-zA-Z_]\w*$', name):
                cols.append(name)
        return cols

    def _split_select_columns(self, text: str) -> list[str]:
        """Split SELECT column list on commas, respecting parentheses."""
        parts = []
        depth = 0
        current = []
        for ch in text:
            if ch == '(':
                depth += 1
                current.append(ch)
            elif ch == ')':
                depth -= 1
                current.append(ch)
            elif ch == ',' and depth == 0:
                parts.append(''.join(current))
                current = []
            else:
                current.append(ch)
        if current:
            parts.append(''.join(current))
        return parts
